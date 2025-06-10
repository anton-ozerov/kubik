import csv
import json
import logging
import os
from pathlib import Path
import re
import time

import pandas as pd
import openpyxl

from data.config import RESULT_DIR

logger = logging.getLogger(f"(Работа с файлами){__name__}")


def clean_price(text: str) -> int:
    """
    Удалит все (включая \\xa0, \\u2009, ₽, обычные пробелы и др.), кроме цифр
    """
    digits_only = re.sub(r"[^\d]", "", text)
    return int(digits_only)


def read_csv_file(file_name: str) -> list[list[str]]:
    res = []
    with open(file_name) as f:
        reader = csv.reader(f)
        for row in reader:
            res.append(row)
    logger.info('Прочитали csv файл')
    return res


def read_xlsx_file(file_name: str):
    sheets_and_columns = {
        0: 16,  # Лист 1 (индекс 0), 17-й столбец (индекс 16)
        1: 17,  # Лист 2, 18-й столбец
        3: 8    # Лист 4, 9-й столбец
    }
    
    wb = openpyxl.load_workbook(file_name, data_only=True)
    result = {}

    for sheet_index, value_column_index in sheets_and_columns.items():
        ws = wb.worksheets[sheet_index]

        # Поиск заголовка "Артикул" в первой строке
        header = [cell.value for cell in ws[1]]
        article_col_index = None
        for idx, title in enumerate(header):
            if title and "Артикул" in str(title):
                article_col_index = idx
                break

        if article_col_index is None:
            print(f"Не найден столбец 'Артикул' на листе {sheet_index + 1}")
            continue

        # Чтение данных
        for row in ws.iter_rows(min_row=2, values_only=True):
            article = row[article_col_index]
            value = row[value_column_index] if value_column_index < len(row) else None
            if article is not None:
                result[article] = value
    return result


def read_json_file(file_name: str) -> list[dict]:
    with open(file_name) as f:
        data = json.load(f)
    logger.info('Прочитали json файл')
    return data


# def find_difference_xlsx_json(xlsx_file_name: str, json_file_name: str, json_from: str = "Детский мир"):
#     xlsx_data = read_csv_file(file_name=xlsx_file_name)
#     json_data = read_json_file(file_name=json_file_name)


def find_difference_csv_json(csv_file_name: str, json_file_name: str, json_from: str = "Детский мир"):
    price_row_num = {
        "Детский мир": int(input("Введите номер столбца детсвого мира: ")),
    }

    csv_data = read_csv_file(file_name=csv_file_name)
    json_data = read_json_file(file_name=json_file_name)

    csv_unic_number_row_num = csv_data[0].index("Артикул")
    # for i in range(len(csv_data[0])):
    #     print(csv_data[0][i], f"i={i}")
    #     print(csv_data[1][i])
    #     print(csv_data[2][i])
    csv_price_row_num = price_row_num[json_from]

    csv_price = {}
    for csv_row in csv_data[1:]:
        csv_uniq_number = csv_row[csv_unic_number_row_num]
        csv_price_item = csv_row[csv_price_row_num]
        csv_price[csv_uniq_number] = csv_price_item
    logger.info('Узнали цены в csv файле')

    json_price = {}
    for obj in json_data:
        json_uniq_number = obj["uniq_number"]
        json_price_item = obj["current_price"]
        json_price[json_uniq_number] = json_price_item
    logger.info('Узнали цены в json файле')

    result = {}
    for uniq_number_item, csv_price_item in csv_price.items():
        if uniq_number_item in json_price:
            json_price_item = json_price[uniq_number_item]
            if not csv_price_item:
                csv_price_item = 0
            json_price_item, csv_price_item = clean_price(str(json_price_item)), clean_price(str(csv_price_item))
            if json_price_item != csv_price_item:
                result[uniq_number_item] = {"csv": csv_price_item, "json": json_price_item}
    logger.info('Нашли различия между csv и json')

    df = pd.DataFrame(
        [
            (uniq_number, prices["csv"], prices["json"])
            for uniq_number, prices in result.items()
        ],
        columns=["Артикул", "Цена в Google Sheet", f"Цена из: {json_from}"],
    )
    
    folder = Path(RESULT_DIR)
    if not folder.exists() or not folder.is_dir():
        os.mkdir(folder)
        logger.info(f'Создана папка {str(folder)}')
    
    file_name = f"RESULT_FOR_{json_from}_{time.time()}"
    path_xlsx_file = folder / f"{file_name}.xlsx"
    
    df.to_excel(path_xlsx_file, index=False)
    logger.info(f'Сохранили различия в {str(path_xlsx_file)}')


# print(read_xlsx_file("results/Аналитика по ценам (2).xlsx"))
